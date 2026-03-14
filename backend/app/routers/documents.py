import asyncio
import csv
import io
import os
import mimetypes
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query, Body
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, cast, Text
from typing import Optional, List
from datetime import date
from ..database import get_db
from ..models.document import Document, ExtractionResult, DocumentStatus, DocumentType
from ..models.user import User, UserRole
from ..schemas.document import DocumentOut, DocumentListOut, ExtractionResultUpdate, AuditLogOut
from ..services.document_service import save_uploaded_file, process_document, add_audit
from ..models.document import AuditLog
from ..schemas.document import DocumentOut, DocumentListOut, ExtractionResultUpdate, AuditLogOut, AuditLogOut
from ..middleware.auth_middleware import get_current_user
from ..config import settings

router = APIRouter(prefix="/documents", tags=["Documents"])

ALLOWED_TYPES = {"pdf", "jpg", "jpeg", "png"}


@router.post("/upload", response_model=DocumentOut, status_code=201)
async def upload_document(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    document_type: DocumentType = DocumentType.invoice,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload a document and trigger background processing."""
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' not allowed. Use PDF, JPG, or PNG.")

    content = await file.read()
    if len(content) > settings.MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Max 10 MB.")

    file_path, stored_name, file_type = save_uploaded_file(content, file.filename)

    document = Document(
        filename=stored_name,
        original_filename=file.filename,
        file_path=file_path,
        file_type=file_type,
        file_size=len(content),
        document_type=document_type,
        owner_id=current_user.id,
    )
    db.add(document)
    db.commit()
    db.refresh(document)

    add_audit(db, document.id, "uploaded", actor_name=current_user.username,
              detail={"filename": file.filename, "file_type": file_type})
    db.commit()

    # Trigger processing pipeline in background
    background_tasks.add_task(process_document, document.id, db)
    return document


@router.get("/", response_model=DocumentListOut)
def list_documents(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    status: Optional[DocumentStatus] = None,
    company: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    sort_by: Optional[str] = Query(None, pattern="^(upload_date|total_amount|company_name|status|original_filename)$"),
    sort_order: Optional[str] = Query("desc", pattern="^(asc|desc)$"),
    tag: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List documents with search, filtering, and sorting. Admins see all, users see their own."""
    query = db.query(Document)

    # Access control
    if current_user.role != UserRole.admin:
        query = query.filter(Document.owner_id == current_user.id)

    # Join with extraction results when needed (search, company filter, amount filter, or sorting by extraction fields)
    needs_join = bool(search or company or min_amount is not None or max_amount is not None or sort_by in ('total_amount', 'company_name'))
    if needs_join:
        query = query.join(ExtractionResult, isouter=True)

    # Filters
    if status:
        query = query.filter(Document.status == status)
    if date_from:
        query = query.filter(Document.upload_date >= date_from)
    if date_to:
        query = query.filter(Document.upload_date <= date_to)
    if search:
        query = query.filter(
            or_(
                Document.original_filename.ilike(f"%{search}%"),
                ExtractionResult.company_name.ilike(f"%{search}%"),
                ExtractionResult.corrected_company_name.ilike(f"%{search}%"),
                ExtractionResult.invoice_number.ilike(f"%{search}%"),
                ExtractionResult.corrected_invoice_number.ilike(f"%{search}%"),
            )
        )
    if company:
        query = query.filter(ExtractionResult.company_name.ilike(f"%{company}%"))
    if min_amount is not None:
        query = query.filter(ExtractionResult.total_amount >= min_amount)
    if max_amount is not None:
        query = query.filter(ExtractionResult.total_amount <= max_amount)
    if tag:
        query = query.filter(cast(Document.tags, Text).ilike(f'%"{tag}"%'))

    # Sorting
    sort_col_map = {
        'upload_date': Document.upload_date,
        'original_filename': Document.original_filename,
        'status': Document.status,
        'total_amount': ExtractionResult.total_amount,
        'company_name': ExtractionResult.company_name,
    }
    col = sort_col_map.get(sort_by or 'upload_date', Document.upload_date)
    query = query.order_by(col.asc() if sort_order == 'asc' else col.desc())

    total = query.count()
    documents = query.offset((page - 1) * page_size).limit(page_size).all()

    return DocumentListOut(total=total, page=page, page_size=page_size, documents=documents)


@router.get("/{document_id}", response_model=DocumentOut)
def get_document(
    document_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get a single document by ID."""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.admin and document.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return document


@router.get("/{document_id}/file")
def get_document_file(
    document_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Serve the original uploaded file."""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.admin and document.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if not os.path.exists(document.file_path):
        raise HTTPException(status_code=404, detail="File not found on disk")

    media_type = mimetypes.guess_type(document.file_path)[0] or 'application/octet-stream'
    return FileResponse(document.file_path, media_type=media_type, filename=document.original_filename)


@router.put("/{document_id}/review", response_model=DocumentOut)
def review_document(
    document_id: int,
    corrections: ExtractionResultUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Submit human corrections for extracted data."""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.admin and document.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    extraction = db.query(ExtractionResult).filter(
        ExtractionResult.document_id == document_id
    ).first()
    if not extraction:
        raise HTTPException(status_code=404, detail="No extraction result found")

    for field, value in corrections.model_dump(exclude_none=True).items():
        setattr(extraction, field, value)
    extraction.is_reviewed = True
    document.status = DocumentStatus.completed
    add_audit(db, document_id, "reviewed", actor_name=current_user.username,
              detail=corrections.model_dump(exclude_none=True))
    db.commit()
    db.refresh(document)
    return document


@router.post("/{document_id}/reprocess", response_model=DocumentOut)
def reprocess_document(
    document_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Re-trigger the processing pipeline for a document."""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.admin and document.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    document.status = DocumentStatus.pending
    document.error_message = None
    add_audit(db, document_id, "reprocessed", actor_name=current_user.username)
    db.commit()
    background_tasks.add_task(process_document, document.id, db)
    return document


@router.get("/export/csv")
def export_documents_csv(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export all documents as a CSV file."""
    query = db.query(Document)
    if current_user.role != UserRole.admin:
        query = query.filter(Document.owner_id == current_user.id)
    documents = query.order_by(Document.upload_date.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'Filename', 'Document Type', 'Status',
        'Invoice Number', 'Company Name', 'Invoice Date', 'Total Amount',
        'Confidence Score', 'Reviewed', 'Upload Date', 'Processed Date',
    ])
    for doc in documents:
        ext = doc.extraction
        writer.writerow([
            doc.id,
            doc.original_filename,
            doc.document_type,
            doc.status,
            ext.corrected_invoice_number or ext.invoice_number if ext else '',
            ext.corrected_company_name or ext.company_name if ext else '',
            ext.corrected_invoice_date or ext.invoice_date if ext else '',
            ext.corrected_total_amount if ext and ext.corrected_total_amount is not None else (ext.total_amount if ext else ''),
            f"{ext.confidence_score:.2f}" if ext and ext.confidence_score else '',
            'Yes' if ext and ext.is_reviewed else 'No',
            doc.upload_date.strftime('%Y-%m-%d %H:%M') if doc.upload_date else '',
            doc.processed_date.strftime('%Y-%m-%d %H:%M') if doc.processed_date else '',
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="documents_export.csv"'},
    )


@router.get("/export/xlsx")
def export_documents_xlsx(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export all documents as a formatted Excel (.xlsx) file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    query = db.query(Document)
    if current_user.role != UserRole.admin:
        query = query.filter(Document.owner_id == current_user.id)
    documents = query.order_by(Document.upload_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    headers = [
        'ID', 'Filename', 'Document Type', 'Status',
        'Invoice Number', 'Company Name', 'Invoice Date', 'Total Amount',
        'Confidence %', 'Reviewed', 'Tags', 'Upload Date', 'Processed Date',
    ]

    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Status colors
    status_fills = {
        'completed':    PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        'failed':       PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        'processing':   PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        'pending':      PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        'review_needed':PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"),
    }

    row_fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    data_align = Alignment(vertical="center")

    for row_idx, doc in enumerate(documents, 2):
        ext = doc.extraction
        row_fill = row_fill_even if row_idx % 2 == 0 else None
        amount = (ext.corrected_total_amount if ext and ext.corrected_total_amount is not None
                  else (ext.total_amount if ext else None))
        tags_str = ', '.join(doc.tags) if doc.tags else ''

        row_data = [
            doc.id,
            doc.original_filename,
            str(doc.document_type).replace('DocumentType.', ''),
            str(doc.status).replace('DocumentStatus.', ''),
            ext.corrected_invoice_number or ext.invoice_number if ext else '',
            ext.corrected_company_name or ext.company_name if ext else '',
            ext.corrected_invoice_date or ext.invoice_date if ext else '',
            amount,
            round(ext.confidence_score * 100, 1) if ext and ext.confidence_score else '',
            'Yes' if ext and ext.is_reviewed else 'No',
            tags_str,
            doc.upload_date.strftime('%Y-%m-%d %H:%M') if doc.upload_date else '',
            doc.processed_date.strftime('%Y-%m-%d %H:%M') if doc.processed_date else '',
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            cell.border = border
            if row_fill:
                cell.fill = row_fill

        # Color status cell
        status_val = str(doc.status).replace('DocumentStatus.', '')
        status_cell = ws.cell(row=row_idx, column=4)
        if status_val in status_fills:
            status_cell.fill = status_fills[status_val]

        # Currency format for amount cell
        amount_cell = ws.cell(row=row_idx, column=8)
        if amount is not None:
            amount_cell.number_format = '"$"#,##0.00'

    # Auto-fit column widths
    col_widths = [6, 35, 14, 14, 18, 24, 14, 14, 12, 10, 20, 18, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="documents_export.xlsx"'},
    )


@router.put("/{document_id}/tags", response_model=DocumentOut)
def update_document_tags(
    document_id: int,
    tags: List[str] = Body(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Set the tags list for a document."""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.admin and document.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    document.tags = [t.strip().lower() for t in tags if t.strip()]
    add_audit(db, document_id, "tags_updated", actor_name=current_user.username,
              detail={"tags": document.tags})
    db.commit()
    db.refresh(document)
    return document


@router.get("/{document_id}/audit", response_model=list[AuditLogOut])
def get_document_audit(
    document_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get the audit log for a document."""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.admin and document.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return db.query(AuditLog).filter(AuditLog.document_id == document_id).order_by(AuditLog.created_at.asc()).all()


@router.delete("/{document_id}", status_code=204)
def delete_document(
    document_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete a document and its extracted data."""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.admin and document.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    import os
    if os.path.exists(document.file_path):
        os.remove(document.file_path)
    db.delete(document)
    db.commit()
